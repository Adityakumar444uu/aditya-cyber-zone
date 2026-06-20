import os
from openpyxl import Workbook, load_workbook
from django.conf import settings

EXCEL_FILE = os.path.join(settings.BASE_DIR, "acz_applications.xlsx")

def sync_application_to_excel(application):
    headers = [
        "ID", "Customer Name", "Aadhaar No", "Contact No",
        "Application Name", "Application No", "Application Date",
        "Status", "Remarks", "Delivery Date", "Last Updated"
    ]

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(headers)

    found_row = None

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == application.id:
            found_row = row
            break

    data = [
        application.id,
        application.customer.name,
        application.customer.aadhaar_no,
        application.customer.contact_no,
        application.application_name,
        application.application_no,
        str(application.application_date),
        application.status,
        application.remarks,
        str(application.delivery_date) if application.delivery_date else "",
        str(application.status_updated_at),
    ]

    if found_row:
        for col, value in enumerate(data, start=1):
            ws.cell(row=found_row, column=col).value = value
    else:
        ws.append(data)

    wb.save(EXCEL_FILE)